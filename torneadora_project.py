import customtkinter
import datetime as dt
from tkcalendar import Calendar, DateEntry
from datetime import date
import pandas as pd
import openpyxl
from openpyxl import Workbook


def change_appearance_mode_event(new_appearance_mode: str):
    customtkinter.set_appearance_mode(new_appearance_mode)


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # main window
        self.comissao_dia = None
        self.calendario = None
        self.calendario2 = None
        self.data = None
        self.botao_calender1 = None
        self.botao_calender2 = None
        self.filter_comb = None
        self.funcionario = None
        self.comissao_ = None
        self.confirm = None
        self.title('Calculador de comissões')
        self.geometry(f'{1100}x{580}')
        self.state('zoomed')

        self.grid_rowconfigure((0, 1, 2), weight=1)

        # Sidebar
        self.sidebar_frame = customtkinter.CTkFrame(self, width=100, corner_radius=5)
        self.sidebar_frame.grid(row=0, column=0, rowspan=5, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(5, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="Opções",
                                                 font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(30, 40))
        self.main_win = customtkinter.CTkButton(self.sidebar_frame, text='Tela \nPrincipal', command=self.tela_inicio)
        self.main_win.grid(row=1, column=0, padx=(20, 30), pady=10)
        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, text='Histórico de\n Comissões',
                                                        command=self.historico)
        self.sidebar_button_1.grid(row=2, column=0, padx=(20, 30), pady=10)
        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, text='Serviços \n Realizados')
        self.sidebar_button_2.grid(row=3, column=0, padx=(20, 30), pady=10)
        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, text='Suporte')
        self.sidebar_button_3.grid(row=4, column=0, padx=(20, 30), pady=10)
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(230, 10))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark",
                                                                                                   "System"],
                                                                       command=change_appearance_mode_event)

        # janela inicial do código
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(0, 10))
        self.label_title = customtkinter.CTkLabel(master=self, text='Cálculo de comissão dos funcionários',
                                                  font=customtkinter.CTkFont(size=30))
        self.label_title.grid(row=0, column=1, padx=(180, 10), pady=10, columnspan=4)
        self.opcao_funcao = customtkinter.CTkComboBox(master=self, values=['Função', 'Torneiro', 'Soldador'])
        self.opcao_funcao.grid(row=1, column=1, padx=(200, 40), pady=0)
        self.opcao_nome = customtkinter.CTkComboBox(master=self, values=['Funcionário', 'Rayner Custódio Souza Ramos',
                                                                         'Jéssica Magre Lemes',
                                                                         'Maria Eduarda Lira Lemes Braz'], )
        self.confirm = customtkinter.CTkButton(self, text='Confirmar', command=self.confirmar)
        self.confirm.grid(row=1, column=3, padx=(40, 40), pady=10)
        self.send = customtkinter.CTkButton(self, text='Enviar', command=self.comissao)
        self.send.grid(row=1, column=4, padx=(10, 40), pady=10)

        # segunda linha
        self.opcao_nome.grid(row=1, column=2, padx=10, pady=10)
        self.valor_servico = customtkinter.CTkEntry(self, placeholder_text='Valor do serviço')
        self.profissao_label = customtkinter.CTkLabel(self, text='')
        self.profissao_label.grid(row=2, column=1, padx=(200, 40), pady=(10, 300))
        self.nome_label = customtkinter.CTkLabel(self, text='')
        self.nome_label.grid(row=2, column=2, padx=10, pady=(10, 300))
        self.valor_comissao = customtkinter.CTkLabel(self, text='')
        self.valor_comissao.grid(row=2, column=4, padx=10, pady=(10, 300))

        # função que retornma para a tela de inicio quando saimos dela

    def tela_inicio(self):
        # grids que deverão ser anulados
        try:
            self.filter_comb.grid_forget()
            self.botao_calender1.grid_forget()
            self.botao_calender2.grid_forget()
            self.calendario.grid_forget()
            self.funcionario.grid_forget()
            self.comissao_dia.grid_forget()
            self.calendario2.grid_forget()
            try:
                self.comissao_dia.grid_forget()
                self.funcionario.grid_forget()
                self.calendario2.grid_forget()
            except:
                self.comissao_dia.configure(text='')
                self.funcionario.configure(text='')
                self.calendario2 = None
        except:
            pass

        self.label_title.configure(text='Cálculo de comissão dos funcionários')
        self.label_title.grid(row=0, column=1, padx=(180, 10), pady=10, columnspan=4)
        self.opcao_funcao.grid(row=1, column=1, padx=(200, 40), pady=0)
        self.confirm.grid(row=1, column=3, padx=(40, 40), pady=10)
        self.opcao_nome.grid(row=1, column=2, padx=10, pady=10)
        self.send.grid(row=1, column=4, padx=(10, 40), pady=10)
        self.profissao_label.grid(row=2, column=1, padx=(200, 40), pady=(10, 300))
        self.nome_label.grid(row=2, column=2, padx=10, pady=(10, 300))
        self.valor_comissao.grid(row=2, column=4, padx=10, pady=(10, 300))
        self.profissao_label.configure(text='')
        self.nome_label.configure(text='')

    # função de confirmar o valor do serviço calcular a comissão e enviar para o banco de dados
    def confirmar(self):
        nome = self.opcao_nome.get()
        funcao = self.opcao_funcao.get()

        if nome == 'Funcionário':
            nome = ""
        if funcao == 'Função':
            funcao = ""
        self.profissao_label.configure(text=f'{funcao}')
        self.nome_label.configure(text=f'{nome}')
        self.valor_servico.grid(row=2, column=3, padx=(40, 40), pady=(10, 300))

    # calculador de valor da 'comissão'
    def comissao(self):
        valor = float(self.valor_servico.get())
        nome = self.opcao_funcao.get()

        percent_t = 0.09
        percent_s = 0.07
        if valor:
            if nome == 'Torneiro':
                self.comissao_ = valor * percent_t
            else:
                if nome == 'Soldador':
                    self.comissao_ = valor * percent_s
                else:
                    self.comissao_ = 'Erro'
        else:
            self.valor_comissao.configure(self, text='Insira um valor, depois clique em enviar.')

        self.valor_comissao.configure(self, text='Comissão = R$ {:.2f}'.format(self.comissao_))

        nome = self.opcao_nome.get()
        funcao = self.opcao_funcao.get()
        # Obter a data atual
        data_atual = date.today()

        # Ler o arquivo Excel existente ou criar um novo se não existir
        try:
            wb = openpyxl.load_workbook('dbtornadora.xlsx')
        except FileNotFoundError:
            wb = Workbook()
            wb.active.append(['Índice', 'Nome', 'Função', 'Data', 'Comissão'])

        # Acessar a planilha ativa
        planilha = wb.active

        # Obter o índice da próxima linha
        indice = planilha.max_row + 1

        # Adicionar os dados à nova linha
        nova_linha = [indice, nome, funcao, data_atual, self.comissao_]
        planilha.append(nova_linha)

        # Salvar o arquivo Excel
        wb.save('dbtornadora.xlsx')

    # histórico que mostrará as comissões que foram dadas no mês
    def historico(self):
        self.label_title.configure(text='Histórico de Comissão')
        self.label_title.grid(padx=(400, 10), pady=50)
        try:
            self.opcao_nome.grid_forget()
            self.opcao_funcao.grid_forget()
            self.confirm.grid_forget()
            self.send.grid_forget()
            self.nome_label.grid_forget()
            self.valor_comissao.grid_forget()
            self.valor_servico.grid_forget()
            self.profissao_label.grid_forget()

        except:
            pass

        self.botao_calender1 = customtkinter.CTkButton(self, text='Data Ínicio: ', command=self.calendario_i)
        self.botao_calender1.grid(row=1, column=1, padx=(200, 40), pady=0)

        self.botao_calender2 = customtkinter.CTkButton(self, text='Data Fim: ', command=self.calendario_f)
        self.botao_calender2.grid(row=1, column=2, padx=(40, 40), pady=0)

        self.filter_comb = customtkinter.CTkComboBox(self, values=['Filtrar por: ', 'Funcionários', 'Todos'])
        self.filter_comb.grid(row=1, column=3, padx=(40, 40), pady=0)

        self.funcionario = customtkinter.CTkLabel(self, text='Funcionário: ')
        self.funcionario.grid(row=2, column=1, padx=(40, 40), pady=(10, 300))

        self.comissao_dia = customtkinter.CTkLabel(self, text='Comissão')
        self.comissao_dia.grid(row=2, column=2, padx=(40, 40), pady=(10, 300))

    def calendario_i(self):
        self.funcionario.grid_forget()
        try:
            self.calendario2.grid_forget()
            self.calendario.grid_forget()
        except:
            pass

        self.calendario = Calendar(self, locale='pt_br')
        self.calendario.grid(row=1, column=1, padx=(300, 50), pady=(10, 240), rowspan=2)

    def calendario_f(self):
        try:
            self.calendario.grid_forget()
            self.calendario2.grid_forget()
        except:
            pass

        self.calendario2 = Calendar(self, locale='pt_br')
        self.calendario2.grid(row=1, column=2, padx=(50, 50), pady=(10, 240), rowspan=2)


if __name__ == "__main__":
    app = App()
    app.mainloop()
